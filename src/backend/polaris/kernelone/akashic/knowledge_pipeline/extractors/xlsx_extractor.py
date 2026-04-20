"""Excel XLSX Extractor for Knowledge Pipeline.

Uses openpyxl for text extraction from .xlsx files.
Provides graceful degradation if openpyxl is not installed.
"""

from __future__ import annotations

import logging
from typing import TYPE_CHECKING

from polaris.kernelone.akashic.knowledge_pipeline.extractors.base import (
    BaseExtractor,
)

logger = logging.getLogger(__name__)

# Try to import openpyxl; if not available, extractor will degrade gracefully
try:
    import openpyxl

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    openpyxl = None  # type: ignore[assignment]

if TYPE_CHECKING:
    from polaris.kernelone.akashic.knowledge_pipeline.extractors.base import (
        ExtractionOptions,
    )
    from polaris.kernelone.akashic.knowledge_pipeline.protocols import (
        DocumentInput,
        ExtractedFragment,
    )


class XlsxExtractor(BaseExtractor):
    """Extractor for Microsoft Excel .xlsx files.

    Extracts cell values as tabular text, preserving sheet structure.

    Supported MIME types:
    - application/vnd.openxmlformats-officedocument.spreadsheetml.sheet
    """

    SUPPORTED_MIME_TYPES = ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",)

    def is_available(self) -> bool:
        """Check if XLSX extraction is available (openpyxl installed)."""
        return OPENPYXL_AVAILABLE

    async def extract(self, doc: DocumentInput) -> list[ExtractedFragment]:  # type: ignore[override]
        """Extract text fragments from an Excel document."""
        from polaris.kernelone.akashic.knowledge_pipeline.protocols import (
            ExtractedFragment,
        )

        if not OPENPYXL_AVAILABLE:
            logger.warning("XLSX extraction unavailable: openpyxl not installed. Install with: pip install openpyxl")
            return self._fallback_extract(doc)

        if isinstance(doc.content, str):
            logger.warning("XLSX content provided as string, not bytes")
            return [
                ExtractedFragment(
                    text=doc.content[:5000],
                    line_start=1,
                    line_end=1,
                    mime_type=doc.mime_type,
                    metadata={"warning": "string_content_not_processed"},
                )
            ]

        try:
            # Run CPU-bound XLSX parsing in thread pool to avoid blocking event loop
            import asyncio
            import io

            loop = asyncio.get_running_loop()

            def _parse_xlsx_bytes() -> list[ExtractedFragment]:
                assert not isinstance(doc.content, str)  # mypy: str | bytes → bytes
                workbook = openpyxl.load_workbook(io.BytesIO(doc.content), data_only=True)
                fragments: list[ExtractedFragment] = []
                global_line = 1

                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]

                    # Sheet header
                    fragments.append(
                        ExtractedFragment(
                            text=f"[SHEET: {sheet_name}]",
                            line_start=global_line,
                            line_end=global_line,
                            mime_type=doc.mime_type,
                            metadata={
                                "sheet": sheet_name,
                                "fragment_type": "xlsx_sheet_header",
                            },
                        )
                    )
                    global_line += 1

                    # Extract rows
                    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                        # Format row as pipe-separated cells
                        cells = [str(cell) if cell is not None else "" for cell in row]
                        if any(cells):  # Skip completely empty rows
                            row_text = " | ".join(cells).strip()
                            fragments.append(
                                ExtractedFragment(
                                    text=row_text,
                                    line_start=global_line,
                                    line_end=global_line,
                                    mime_type=doc.mime_type,
                                    metadata={
                                        "sheet": sheet_name,
                                        "row": row_idx,
                                        "fragment_type": "xlsx_row",
                                    },
                                )
                            )
                            global_line += 1

                logger.debug(
                    "Extracted %d fragments from XLSX (%d sheets)",
                    len(fragments),
                    len(workbook.sheetnames),
                )
                return fragments

            return await loop.run_in_executor(None, _parse_xlsx_bytes)

        except (RuntimeError, ValueError) as exc:
            logger.warning("XLSX extraction failed: %s", exc)
            return self._fallback_extract(doc)

    def _fallback_extract(self, doc: DocumentInput) -> list[ExtractedFragment]:
        """Fallback when openpyxl is unavailable."""
        from polaris.kernelone.akashic.knowledge_pipeline.protocols import (
            ExtractedFragment,
        )

        if isinstance(doc.content, str):
            return [
                ExtractedFragment(
                    text=doc.content[:5000],
                    line_start=1,
                    line_end=1,
                    mime_type=doc.mime_type,
                    metadata={"fallback": "string_content"},
                )
            ]

        return [
            ExtractedFragment(
                text="[XLSX content could not be extracted - openpyxl not installed]",
                line_start=1,
                line_end=1,
                mime_type=doc.mime_type,
                metadata={"error": "openpyxl_missing"},
            )
        ]

    def _do_extract(  # type: ignore[override]
        self,
        text: str,
        options: ExtractionOptions,
    ) -> list[ExtractedFragment]:
        """Not used - override extract() handles bytes directly."""
        return []


__all__ = ["XlsxExtractor"]
